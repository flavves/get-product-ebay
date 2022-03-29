# -*- coding: utf-8 -*-
"""
Created on Sat Mar 26 15:53:49 2022

@author: okmen
"""


import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import selenium
from  selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import requests
#import xlsxwriter

version="0"




try:
        
    version_bulma=requests.get('https://raw.githubusercontent.com/flavves/TestOkur/main/version.txt')
    version=version_bulma.text
except:
    pass




if version=="1.1\n":
    
    dosya=open("butunlinker.txt", 'r')
    
    linkokudum=dosya.readline()
    
    linkokudum=linkokudum.split(";")
    
    
    driver = webdriver.Chrome(executable_path=r'chromedriver.exe')
    
    
    sira=4
    saticistokkodu="PSPS"
    barkod="BPSPS"
    varyantKodu="VPSPS0"
    kitap = openpyxl.load_workbook("taslak.xlsx")
    sayfa = kitap.get_sheet_by_name("Spor Paspaslar")
    
      
    
    driver.get("https://www.ebay.co.uk/")
    time.sleep(10)
    driver.find_element_by_xpath('//*[@id="gdpr-banner-accept"]').click()
    kacinci_link=0
    for linkgeldi in linkokudum:
        kacinci_link=kacinci_link+1
        print("%s sırada ve %s kadar link var"%(kacinci_link,str(len(linkokudum))))  
        donguden_cik=0
        try:
            
            driver.get(linkgeldi)
        except:
            kitap.save("cikti.xlsx")
            kitap.close()
            break
            
        
    
        
        
        
        
        
        
        TrimColorVaryant=""
        CarpetColourVaryant=""
        CarpetGrade=""
        varyasyonisim1=""
        varyasyonisim2=""
        varyasyonisim3=""
        
        #varyasyon sayısını ogrenme satırı
        
        #############################
        
        kactanevaryasyonvarogren=0
        
        varyasyon1uzunluk=0
        varyasyon2uzunluk=0
        varyasyon3uzunluk=0
        
        
        #sample urun varsa dışarıda bırakıyorum sonra onlara bakacağım anlamına geliyor
        for sample_bul in range(0,300):
            try:
                    
                sample_urun_tespit=driver.find_element_by_xpath('//*[@id="msku-opt-'+str(sample_bul)+'"]').text
                if sample_urun_tespit=="Sample" or sample_urun_tespit=="sample":
                    print("buldum")
                    dosya=open("samplelilinkler.txt", 'a')
                    dosya.write(linkgeldi+";")
                    donguden_cik=1
                    break
            except:
                pass
        
        if donguden_cik==0:
            
            try:
            
                try:
                    varyasyonisim1=driver.find_element_by_xpath('//*[@id="msku-sel-1-lbl"]').text
                    kactanevaryasyonvarogren+=1
                    
                    varyasyon1_elemalari=driver.find_element_by_xpath('//*[@id="msku-sel-1"]').text.split("\n")
                    varyasyon1_elemalari.pop(0)    
                    varyasyon1uzunluk=len(varyasyon1_elemalari)
                    
                    
                except:
                    kactanevaryasyonvarogren=0
                try:
                    varyasyonisim2=driver.find_element_by_xpath('//*[@id="msku-sel-2-lbl"]').text
                    kactanevaryasyonvarogren+=1
                    
                    varyasyon2_elemalari=driver.find_element_by_xpath('//*[@id="msku-sel-2"]').text.split("\n")
                    varyasyon2_elemalari.pop(0)    
                    varyasyon2uzunluk=len(varyasyon2_elemalari)
                    
                    
                except:pass
                try:
                    varyasyonisim3=driver.find_element_by_xpath('//*[@id="msku-sel-3-lbl"]').text
                    kactanevaryasyonvarogren+=1
                    
                    varyasyon3_elemalari=driver.find_element_by_xpath('//*[@id="msku-sel-3"]').text.split("\n")
                    varyasyon3_elemalari.pop(0)    
                    varyasyon3uzunluk=len(varyasyon3_elemalari)
                    
                    
                except:pass
                
                ############################
                
                #açıklama için satır
                
                #############################
                
                
                    
                try:
                    xpath_aciklama="/html/body/div[5]/div[5]/div[1]/div[5]/div[2]/div/div[6]/div/div[7]/div[3]/div/span[2]/span["+str(2)+"]/a"      
                    driver.find_element_by_xpath(xpath_aciklama).click()
                        
                except:
                    pass
                
                
                
                aciklama=[]
                for aciklama_icin in range(2,20):
                
                    try:
                        
                        xpath_aciklama="/html/body/div[5]/div[5]/div[1]/div[5]/div[2]/div/div[6]/div/div[7]/div[3]/div/span[2]/span["+str(aciklama_icin)+"]/a"      
                        driver.find_element_by_xpath(xpath_aciklama).click()
                        
                        try:
                
                             element = WebDriverWait(driver, 10).until(
                                 EC.visibility_of_element_located((By.XPATH, "/html/body/div[5]/div[5]/div[1]/div[5]/div[2]/div/div[6]/div/div[7]/table/tbody"))
                             )
                        except:
                            pass
                        aciklama.append(driver.find_element_by_xpath("/html/body/div[5]/div[5]/div[1]/div[5]/div[2]/div/div[6]/div/div[7]/table/tbody").text)
                    except Exception as e:
                        print(e)
                        break
                
                    
                aciklamabitti=""
                for aciklamabirlestir in aciklama:
                    aciklamabitti=aciklamabitti+"\n"+aciklamabirlestir
                uyumlu_marka=""
                uyumlu_model=""
                
                try:
                        
                    uyumlu_marka=aciklamabitti.split("\n")[2].split(" ")[2]
                    uyumlu_model=aciklamabitti.split("\n")[2].split(" ")[3]
                except:
                    pass
                
                ############################
                
                # resimleri çekiyorum
                resimler=[]
                #############################
                for resimcik in range(0,6):
                    driver.execute_script("scrollBy(0,-1500);")
                    try:
                            
                        xpath_kucukresim='//*[@id="vi_main_img_fs_thImg'+str(resimcik)+'"]/div/img'
                        driver.find_element_by_xpath(xpath_kucukresim).click()
                        
                        try:
                             element = WebDriverWait(driver, 10).until(
                                 EC.visibility_of_element_located((By.XPATH, '//*[@id="icImg"]'))
                             )
                        except:
                            pass
                            
                        
                        resimler.append(driver.find_element_by_xpath('//*[@id="icImg"]').get_attribute('src'))
                
                        
                    except:
                        break
                
                
                # ad çektim
                ad=driver.find_element_by_xpath('//*[@id="LeftSummaryPanel"]/div[1]/div[1]/div/h1/span').text
                
                
                
                #varyantKodu="VPSPS"+str(int(varyantKodu.split("VPSPS")[1])+1)
                varyantKodu=driver.find_element_by_xpath('//*[@id="descItemNumber"]').text
            
                
                
                
                #varyantları sıralama
                
                variant_1_Ad_yazilacak=""
                variant_2_Ad_yazilacak=""
                variant_3_Ad_yazilacak=""
                variant_4_Ad_yazilacak=""
                variant_1_Ad=""
                variant_2_Ad=""
                variant_3_Ad=""
                
                
                #varyasyon 0 ise yapılacaklar
                
                
                if kactanevaryasyonvarogren==0:
                
                            try:
                                  time.sleep(0.1)
                                  element = WebDriverWait(driver, 10).until(
                                      EC.visibility_of_element_located((By.XPATH, '//*[@id="prcIsum"]'))
                                  )
                            except:
                                 pass
                             
                             
                            fiyat=driver.find_element_by_xpath('//*[@id="prcIsum"]').get_attribute('content')
                    
                            barkod="BPSPS"+str(sira)
                            saticistokkodu="PSPS"+str(sira)
                    
                             
                             
                             
                            #excele ürün ekleme
                            
                            ##########################################################
                            sayfa.cell(row=sira,column=1,value=ad)
                            sayfa.cell(row=sira,column=2,value=saticistokkodu)
                            sayfa.cell(row=sira,column=3,value=barkod)
                            sayfa.cell(row=sira,column=4,value=varyantKodu)
                    
                            sayfa.cell(row=sira,column=5,value=aciklamabitti)
                            #hazır olanlar
                            sayfa.cell(row=sira,column=6,value="Tegin")
                            sayfa.cell(row=sira,column=7,value="3")
                            sayfa.cell(row=sira,column=8,value="8") 
                            sayfa.cell(row=sira,column=9,value="24")
                            
                            #resimler
                            try:sayfa.cell(row=sira,column=10,value=resimler[0]) 
                            except:pass
                            try:sayfa.cell(row=sira,column=11,value=resimler[1]) 
                            except:pass
                            try:sayfa.cell(row=sira,column=12,value=resimler[2])
                            except:pass
                            try:sayfa.cell(row=sira,column=13,value=resimler[3])
                            except:pass
                            try:sayfa.cell(row=sira,column=14,value=resimler[4]) 
                            except:pass
                            
                            sayfa.cell(row=sira,column=15,value=fiyat)
                            
                            sayfa.cell(row=sira,column=16,value="10000")
                            
                           
                            
                            
                            #sayfa.cell(row=sira,column=17,value="NAN")
                
                            
                            sayfa.cell(row=sira,column=21,value=uyumlu_marka)
                            sayfa.cell(row=sira,column=22,value=uyumlu_model)
                            
                            sira=sira+1
                            
                    
                             
                         
                         
                
                
            
                
                
                
                
                
                #varyasyon 1 ise yapılacaklar
                
                
                if kactanevaryasyonvarogren==1:
                    
                
                
                    for varyasyon1_sayac in range(0,varyasyon1uzunluk):
                        
                            
                            try:
                                  
                                  
                                  driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon1_sayac)+'"]').click()
                                  variant_1_Ad=driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon1_sayac)+'"]').text
                        
                               
                            except:
                                  pass
                       
                    
                    
                            try:
                                  time.sleep(0.1)
                                  element = WebDriverWait(driver, 10).until(
                                      EC.visibility_of_element_located((By.XPATH, '//*[@id="prcIsum"]'))
                                  )
                            except:
                                 pass
                             
                             
                            fiyat=driver.find_element_by_xpath('//*[@id="prcIsum"]').get_attribute('content')
                    
                            barkod="BPSPS"+str(sira)
                            saticistokkodu="PSPS"+str(sira)
                    
                             
                             
                             
                            #excele ürün ekleme
                            
                            ##########################################################
                            sayfa.cell(row=sira,column=1,value=ad)
                            sayfa.cell(row=sira,column=2,value=saticistokkodu)
                            sayfa.cell(row=sira,column=3,value=barkod)
                            sayfa.cell(row=sira,column=4,value=varyantKodu)
                    
                            sayfa.cell(row=sira,column=5,value=aciklamabitti)
                            #hazır olanlar
                            sayfa.cell(row=sira,column=6,value="Tegin")
                            sayfa.cell(row=sira,column=7,value="3")
                            sayfa.cell(row=sira,column=8,value="8") 
                            sayfa.cell(row=sira,column=9,value="24")
                            
                            #resimler
                            try:sayfa.cell(row=sira,column=10,value=resimler[0]) 
                            except:pass
                            try:sayfa.cell(row=sira,column=11,value=resimler[1]) 
                            except:pass
                            try:sayfa.cell(row=sira,column=12,value=resimler[2])
                            except:pass
                            try:sayfa.cell(row=sira,column=13,value=resimler[3])
                            except:pass
                            try:sayfa.cell(row=sira,column=14,value=resimler[4]) 
                            except:pass
                            
                            sayfa.cell(row=sira,column=15,value=fiyat)
                            
                            sayfa.cell(row=sira,column=16,value="10000")
                            
                            ########################
            
                            
                            sayfa.cell(row=sira,column=17,value=varyasyonisim1+variant_1_Ad)          
                            sayfa.cell(row=sira,column=20,value=uyumlu_marka)
                            sayfa.cell(row=sira,column=21,value=uyumlu_model)
                            
                            sira=sira+1
                            
                    
                             
                         
                         
                
                
                
            
                
                
                #varyasyon 2 ise yapılacaklar
                
                
                if kactanevaryasyonvarogren==2:
                    
                
                
                    for varyasyon1_sayac in range(0,varyasyon1uzunluk):
                        for varyasyon2_sayac in range(varyasyon1uzunluk,varyasyon2uzunluk+varyasyon1uzunluk):
                            
                            try:
                                  
                                  
                                  driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon1_sayac)+'"]').click()
                                  variant_1_Ad=driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon1_sayac)+'"]').text
                        
                                  driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon2_sayac)+'"]').click()
                                  variant_2_Ad=driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon2_sayac)+'"]').text
                                  
                            except:
                                  pass
                       
                    
                    
                            try:
                                  time.sleep(0.1)
                                  element = WebDriverWait(driver, 10).until(
                                      EC.visibility_of_element_located((By.XPATH, '//*[@id="prcIsum"]'))
                                  )
                            except:
                                 pass
                             
                             
                            fiyat=driver.find_element_by_xpath('//*[@id="prcIsum"]').get_attribute('content')
                    
                            barkod="BPSPS"+str(sira)
                            saticistokkodu="PSPS"+str(sira)
                    
                             
                             
                             
                            #excele ürün ekleme
                            
                            ##########################################################
                            sayfa.cell(row=sira,column=1,value=ad)
                            sayfa.cell(row=sira,column=2,value=saticistokkodu)
                            sayfa.cell(row=sira,column=3,value=barkod)
                            sayfa.cell(row=sira,column=4,value=varyantKodu)
                    
                            sayfa.cell(row=sira,column=5,value=aciklamabitti)
                            #hazır olanlar
                            sayfa.cell(row=sira,column=6,value="Tegin")
                            sayfa.cell(row=sira,column=7,value="3")
                            sayfa.cell(row=sira,column=8,value="8") 
                            sayfa.cell(row=sira,column=9,value="24")
                            
                            #resimler
                            try:sayfa.cell(row=sira,column=10,value=resimler[0]) 
                            except:pass
                            try:sayfa.cell(row=sira,column=11,value=resimler[1]) 
                            except:pass
                            try:sayfa.cell(row=sira,column=12,value=resimler[2])
                            except:pass
                            try:sayfa.cell(row=sira,column=13,value=resimler[3])
                            except:pass
                            try:sayfa.cell(row=sira,column=14,value=resimler[4]) 
                            except:pass
                            
                            sayfa.cell(row=sira,column=15,value=fiyat)
                            
                            sayfa.cell(row=sira,column=16,value="10000")
                            
                            
                            
                            
                            sayfa.cell(row=sira,column=17,value=varyasyonisim1+variant_1_Ad)  
                            sayfa.cell(row=sira,column=18,value=varyasyonisim2+variant_2_Ad)
                            sayfa.cell(row=sira,column=20,value=uyumlu_marka)
                            sayfa.cell(row=sira,column=21,value=uyumlu_model)
                            
                            sira=sira+1
                            
                    
                             
                         
                         
                
                
                
                
                
                
                
                
                
                
                #varyasyon 3 ise yapılacaklar
                
                
                if kactanevaryasyonvarogren==3:
                        
                    for varyasyon1_sayac in range(0,varyasyon1uzunluk):
                        for varyasyon2_sayac in range(varyasyon1uzunluk,varyasyon2uzunluk+varyasyon1uzunluk):
                            for varyasyon3_sayac in range(varyasyon2uzunluk+varyasyon1uzunluk,varyasyon3uzunluk+varyasyon2uzunluk+varyasyon1uzunluk):
                                
                                
                                
                                try:
                                    
                                    #driver.find_element_by_xpath('//*[@id="msku-sel-1"]').click()
                                    driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon1_sayac)+'"]').click()
                                    variant_1_Ad=driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon1_sayac)+'"]').text
                    
                    
                    
                                    
                                    #driver.find_element_by_xpath('//*[@id="msku-sel-2"]').click()
                                    driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon2_sayac)+'"]').click()
                                    variant_2_Ad=driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon2_sayac)+'"]').text
                                    
                                    
                                    #driver.find_element_by_xpath('//*[@id="msku-sel-3"]').click()
                                    driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon3_sayac)+'"]').click()
                                    variant_3_Ad=driver.find_element_by_xpath('//*[@id="msku-opt-'+str(varyasyon3_sayac)+'"]').text
                                    
                                except:
                                    pass
                                
                                
                                
                                try:
                                     time.sleep(0.1)
                                     element = WebDriverWait(driver, 10).until(
                                         EC.visibility_of_element_located((By.XPATH, '//*[@id="prcIsum"]'))
                                     )
                                except:
                                    pass
                                
                                
                                fiyat=driver.find_element_by_xpath('//*[@id="prcIsum"]').get_attribute('content')
                    
                    
                    
                                
                                
                                
                    
                                barkod="BPSPS"+str(sira)
                                saticistokkodu="PSPS"+str(sira)
                                
                                
                                
                                
                                #excele ürün ekleme
                                
                                ##########################################################
                                sayfa.cell(row=sira,column=1,value=ad)
                                sayfa.cell(row=sira,column=2,value=saticistokkodu)
                                sayfa.cell(row=sira,column=3,value=barkod)
                                sayfa.cell(row=sira,column=4,value=varyantKodu)
                        
                                sayfa.cell(row=sira,column=5,value=aciklamabitti)
                                #hazır olanlar
                                sayfa.cell(row=sira,column=6,value="Tegin")
                                sayfa.cell(row=sira,column=7,value="3")
                                sayfa.cell(row=sira,column=8,value="8") 
                                sayfa.cell(row=sira,column=9,value="24")
                                
                                #resimler
                                try:sayfa.cell(row=sira,column=10,value=resimler[0]) 
                                except:pass
                                try:sayfa.cell(row=sira,column=11,value=resimler[1]) 
                                except:pass
                                try:sayfa.cell(row=sira,column=12,value=resimler[2])
                                except:pass
                                try:sayfa.cell(row=sira,column=13,value=resimler[3])
                                except:pass
                                try:sayfa.cell(row=sira,column=14,value=resimler[4]) 
                                except:pass
                                
                                sayfa.cell(row=sira,column=15,value=fiyat)
                                
                                sayfa.cell(row=sira,column=16,value="10000")
                                
            
                                sayfa.cell(row=sira,column=17,value=varyasyonisim1+variant_1_Ad)
                                sayfa.cell(row=sira,column=18,value=varyasyonisim2+variant_2_Ad)
                                sayfa.cell(row=sira,column=19,value=varyasyonisim3+variant_3_Ad)
                                sayfa.cell(row=sira,column=20,value=uyumlu_marka)
                                sayfa.cell(row=sira,column=21,value=uyumlu_model)
                                
                                sira=sira+1
                                
        
                ############################
            except:
                kitap.save("cikti.xlsx")
                kitap.close()
            
            try:
                      
                kitap.save("cikti.xlsx")
                kitap.close()        
            except:
                pass  
            
    dosya.close()   
    
try:
          
    kitap.save("cikti.xlsx")
    kitap.close()        
except:
    pass           
        
        
        
        
    
    
    
    
    
    








